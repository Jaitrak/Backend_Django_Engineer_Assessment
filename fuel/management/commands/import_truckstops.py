import os
import csv
import json
import logging
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Avg
from fuel.models import TruckStop

logger = logging.getLogger("fuel")


class Command(BaseCommand):
    help = "Imports truckstop fuel prices from a CSV or Excel file and geocodes them using local mappings."

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to the Excel or CSV file containing fuel prices.",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]

        if not os.path.exists(file_path):
            raise CommandError(f"File not found at path: {file_path}")

        # 1. Load Local Geocoded Cities Database
        base_dir = os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        )
        mapping_path = os.path.join(base_dir, "data", "geocoded_cities.json")
        if not os.path.exists(mapping_path):
            raise CommandError(
                f"Local geocoded cities mapping database not found at {mapping_path}. "
                "Please ensure geocoded_cities.json is committed in fuel/data/."
            )

        with open(mapping_path, "r", encoding="utf-8") as f:
            city_coordinates = json.load(f)

        self.stdout.write("Reading input file...")
        raw_rows = []
        is_excel = file_path.endswith((".xlsx", ".xls"))

        if is_excel:
            try:
                import openpyxl

                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet = wb.active
                iter_rows = sheet.iter_rows(values_only=True)
                header_row = next(iter_rows, None)
                if not header_row:
                    raise CommandError("The Excel file is empty.")

                headers = [str(h).strip() if h is not None else "" for h in header_row]
                for row in iter_rows:
                    if not any(row):
                        continue
                    row_dict = {
                        headers[i]: row[i] for i in range(min(len(headers), len(row)))
                    }
                    raw_rows.append(row_dict)
            except ImportError:
                raise CommandError(
                    "openpyxl is required to read Excel files. Run: pip install openpyxl"
                )
            except Exception as e:
                raise CommandError(f"Error reading Excel file: {str(e)}")
        else:
            try:
                with open(file_path, mode="r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        raw_rows.append(row)
            except Exception as e:
                raise CommandError(f"Error reading CSV file: {str(e)}")

        total_processed = len(raw_rows)
        self.stdout.write(f"Found {total_processed} raw records. Processing...")

        # 2. Clean and Deduplicate (Keep cheapest price for duplicate OPIS Truckstop IDs)
        cleaned_stops = {}
        duplicate_count = 0

        for row in raw_rows:
            try:
                raw_tid = row.get("OPIS Truckstop ID")
                if raw_tid is None:
                    continue
                # Convert float/int strings to exact ID string
                truckstop_id = str(int(float(raw_tid))).strip()

                name = str(row.get("Truckstop Name", "")).strip()
                address = str(row.get("Address", "")).strip()
                city = str(row.get("City", "")).strip().lower()
                state = str(row.get("State", "")).strip().upper()

                raw_price = row.get("Retail Price")
                if raw_price is None:
                    continue
                retail_price = float(raw_price)

                if not truckstop_id or not name or not address or not city or not state:
                    continue

                if truckstop_id in cleaned_stops:
                    duplicate_count += 1
                    # Keep the one with the cheaper retail price
                    if retail_price < cleaned_stops[truckstop_id]["retail_price"]:
                        cleaned_stops[truckstop_id] = {
                            "name": name,
                            "address": address,
                            "city": city,
                            "state": state,
                            "retail_price": retail_price,
                        }
                else:
                    cleaned_stops[truckstop_id] = {
                        "name": name,
                        "address": address,
                        "city": city,
                        "state": state,
                        "retail_price": retail_price,
                    }
            except (ValueError, TypeError):
                continue

        # 3. Geocode and Prepare Models
        db_objects = []
        failed_count = 0
        success_count = 0

        for tid, data in cleaned_stops.items():
            city_key = f"{data['city']},{data['state']}"
            if city_key in city_coordinates:
                coords = city_coordinates[city_key]
                stop_obj = TruckStop(
                    truckstop_id=tid,
                    name=data["name"],
                    address=data["address"],
                    city=data["city"].title(),
                    state=data["state"],
                    retail_price=data["retail_price"],
                    latitude=coords["latitude"],
                    longitude=coords["longitude"],
                )
                db_objects.append(stop_obj)
                success_count += 1
            else:
                failed_count += 1
                logger.warning(
                    f"Unmapped location skipped: {data['city'].title()}, {data['state']} (ID: {tid})"
                )

        # 4. Clear Database & Bulk Insert
        self.stdout.write("Clearing existing database entries...")
        with transaction.atomic():
            TruckStop.objects.all().delete()
            self.stdout.write(f"Inserting {len(db_objects)} geocoded records...")
            # Insert in batches of 500
            TruckStop.objects.bulk_create(db_objects, batch_size=500)

        # 5. Calculate Average Imported Price
        avg_price = TruckStop.objects.aggregate(Avg("retail_price"))[
            "retail_price__avg"
        ]
        avg_price_formatted = f"${avg_price:.2f}" if avg_price is not None else "N/A"

        # 6. Print Summary Statistics
        self.stdout.write(self.style.SUCCESS("\n--- Import Summary Statistics ---"))
        self.stdout.write(f"Imported: {total_processed}")
        self.stdout.write(f"Duplicates removed: {duplicate_count}")
        self.stdout.write(f"Geocoded: {success_count}")
        self.stdout.write(f"Failed: {failed_count}")
        self.stdout.write(f"Average fuel price: {avg_price_formatted}")
        self.stdout.write(self.style.SUCCESS("---------------------------------\n"))
